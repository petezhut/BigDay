from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from database import db_method
import datetime
import os

@db_method
def getAttending(**kwargs):
    db = kwargs.pop('db')
    return db.rsvp.find({'attending' : kwargs.get('attending', 'Yes')})

class RSVPBackup:
    def __init__(self):
        self.wb = Workbook()
        self.attending()
        self.not_attending()

    def __del__(self):
        now = datetime.datetime.now()
        dateref = str(now.date()).replace("-", "_")
        if not os.path.isdir(dateref):
            os.mkdir(dateref)
        fname = "%s/backup.xlsx" % (dateref)
        self.wb.save(filename = fname)

    def not_attending(self):
        self.wb.create_sheet(1)
        self.notattending = self.wb.worksheets[1]
        self.notattending.title = 'Not Attending'
        self.notattending.cell("A1").value = 'RSVP Name'
        ref = 2
        for count, data in enumerate(getAttending(attending = 'No')):
            self.notattending.cell("A%d" % (ref)).value = "%s %s" % (data['rsvp_fname'][0], data['rsvp_lname'][0])

    def attending(self):
        self.attending = self.wb.worksheets[0]
        self.attending.title = 'Attending'
        self.attending.cell("A1").value = 'RSVP Name'
        self.attending.cell("B1").value = 'Guest Name'
        self.attending.cell("C1").value = 'Guest Meal'
        self.attending.cell("D1").value = "Child's Age"
        ref = 2
        for count, data in enumerate(getAttending()):
            rsvp_name = "%s %s" % (data['rsvp_fname'][0], data['rsvp_lname'][0])
            for a in range(int(data['count'][0])):
                self.attending.cell("A%d" % (ref)).value = rsvp_name
                self.attending.cell("B%d" % (ref)).value = data['guest_%d_name' % (a)][0]
                self.attending.cell("C%d" % (ref)).value = data['guest_%d_meal' % (a)][0]
                if data.has_key('guest_%d_age' % (a)):
                    self.attending.cell("D%d" % (ref)).value = data['guest_%d_age' % (a)][0]
                ref += 1

if __name__ == '__main__':
    RSVPBackup()
